"""
Update OQI Impact Workbook with SDG Interlinkages
Based on the research paper on SDG interlinkages (s11625-023-01336-x.pdf)
"""

import openpyxl
from openpyxl.styles import Font, Alignment
import sys
import io

# Set UTF-8 encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Load the workbook
wb_path = r"d:\Projects\OQI-UC002-DWave\Proposal\OQI_Impact_Workbook.xlsx"
wb = openpyxl.load_workbook(wb_path)
ws = wb['Impact Rows']

# Define SDG interlinkages based on the food production optimization context
# and the research paper findings
sdg_interlinkages = {
    'SDG 2': {
        'positive': [
            ('SDG 3', 'Better nutrition improves health outcomes & reduces disease burden'),
            ('SDG 1', 'Increased agricultural productivity reduces poverty in rural areas'),
            ('SDG 8', 'Sustainable agriculture creates decent work opportunities'),
            ('SDG 12', 'Optimized crop planning reduces food waste & resource use'),
        ],
        'negative': [
            ('SDG 13', 'Intensive agriculture may increase GHG emissions without optimization'),
            ('SDG 15', 'Land conversion for agriculture can harm terrestrial ecosystems'),
            ('SDG 6', 'Agricultural expansion may strain water resources & quality'),
            ('SDG 14', 'Agricultural runoff can negatively impact aquatic ecosystems'),
        ]
    },
    'SDG 3': {
        'positive': [
            ('SDG 2', 'Improved health enables better agricultural productivity'),
            ('SDG 4', 'Healthy populations have better educational outcomes'),
            ('SDG 8', 'Healthier workforce supports economic growth'),
        ],
        'negative': [
            ('SDG 13', 'Healthcare infrastructure may increase energy use & emissions'),
            ('SDG 12', 'Medical/health activities can generate significant waste'),
        ]
    },
    'SDG 9': {
        'positive': [
            ('SDG 8', 'Innovation infrastructure drives economic growth & jobs'),
            ('SDG 11', 'Digital platforms enable smart city & sustainable planning'),
            ('SDG 2', 'Technology platforms improve agricultural efficiency & yields'),
            ('SDG 17', 'Digital tools strengthen partnerships & data sharing'),
        ],
        'negative': [
            ('SDG 13', 'Technology production & operation increase energy demand'),
            ('SDG 12', 'Electronic waste from technology infrastructure'),
            ('SDG 10', 'Digital divide may worsen inequalities without access policies'),
        ]
    },
    'SDG 13': {
        'positive': [
            ('SDG 7', 'Climate action drives renewable energy adoption'),
            ('SDG 15', 'Climate mitigation protects terrestrial ecosystems'),
            ('SDG 14', 'Reduced emissions benefit ocean health'),
        ],
        'negative': [
            ('SDG 1', 'Climate policies may increase costs for poor households'),
            ('SDG 2', 'Climate mitigation constraints may limit agricultural expansion'),
            ('SDG 8', 'Transition costs may temporarily slow economic growth'),
        ]
    }
}

# Find the last column with data
max_col = ws.max_column
print(f"Current max column: {max_col}")

# Add new column headers
positive_col = max_col + 1
negative_col = max_col + 2

ws.cell(row=1, column=positive_col, value="Positive SDG Interlinkages")
ws.cell(row=1, column=negative_col, value="Negative SDG Interlinkages")

# Style the headers
for col in [positive_col, negative_col]:
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Process each data row
print("\nProcessing rows:")
for row_idx in range(2, ws.max_row + 1):
    sdg_cell = ws.cell(row=row_idx, column=7)  # Column G (SDG)
    sdg_value = sdg_cell.value
    
    if sdg_value:
        print(f"\nRow {row_idx}: {sdg_value}")
        
        # Extract SDG number
        sdg_key = None
        if '2' in str(sdg_value) and 'Zero Hunger' in str(sdg_value):
            sdg_key = 'SDG 2'
        elif '3' in str(sdg_value) and ('Health' in str(sdg_value) or 'health' in str(sdg_value)):
            sdg_key = 'SDG 3'
        elif '9' in str(sdg_value) and ('Innovation' in str(sdg_value) or 'Infrastructure' in str(sdg_value)):
            sdg_key = 'SDG 9'
        elif '13' in str(sdg_value) and ('Climate' in str(sdg_value) or 'climate' in str(sdg_value)):
            sdg_key = 'SDG 13'
        
        if sdg_key and sdg_key in sdg_interlinkages:
            # Add positive interlinkages
            positive_links = sdg_interlinkages[sdg_key]['positive']
            positive_text = "; ".join([f"{sdg}: {desc}" for sdg, desc in positive_links])
            
            # Truncate to 95 characters per cell as specified
            if len(positive_text) > 95:
                # Try to fit as many complete entries as possible
                truncated = []
                current_len = 0
                for sdg, desc in positive_links:
                    entry = f"{sdg}: {desc}"
                    if current_len + len(entry) + 2 <= 95:  # +2 for "; "
                        truncated.append(entry)
                        current_len += len(entry) + 2
                    else:
                        break
                positive_text = "; ".join(truncated)
            
            ws.cell(row=row_idx, column=positive_col, value=positive_text)
            ws.cell(row=row_idx, column=positive_col).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            print(f"  Positive: {positive_text[:80]}...")
            
            # Add negative interlinkages
            negative_links = sdg_interlinkages[sdg_key]['negative']
            negative_text = "; ".join([f"{sdg}: {desc}" for sdg, desc in negative_links])
            
            # Truncate to 95 characters per cell as specified
            if len(negative_text) > 95:
                truncated = []
                current_len = 0
                for sdg, desc in negative_links:
                    entry = f"{sdg}: {desc}"
                    if current_len + len(entry) + 2 <= 95:
                        truncated.append(entry)
                        current_len += len(entry) + 2
                    else:
                        break
                negative_text = "; ".join(truncated)
            
            ws.cell(row=row_idx, column=negative_col, value=negative_text)
            ws.cell(row=row_idx, column=negative_col).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            print(f"  Negative: {negative_text[:80]}...")

# Set column widths
ws.column_dimensions[openpyxl.utils.get_column_letter(positive_col)].width = 50
ws.column_dimensions[openpyxl.utils.get_column_letter(negative_col)].width = 50

# Save the workbook
wb.save(wb_path)
print(f"\n\nWorkbook updated successfully!")
print(f"Added columns: {openpyxl.utils.get_column_letter(positive_col)} and {openpyxl.utils.get_column_letter(negative_col)}")

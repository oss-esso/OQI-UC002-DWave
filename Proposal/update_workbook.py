"""
Script to read and update the OQI Impact Workbook with SDG interlinkages
"""
from openpyxl import load_workbook
import os

# Load the workbook
workbook_path = r"d:\Projects\OQI-UC002-DWave\Proposal\OQI_Impact_Workbook.xlsx"
wb = load_workbook(workbook_path)

# Print all sheet names
print("Available sheets:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

# Let's examine each sheet
for sheet_name in wb.sheetnames:
    print(f"\n=== Sheet: {sheet_name} ===")
    ws = wb[sheet_name]
    
    # Print first 20 rows to understand structure
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 20:
            break
        print(f"Row {i}: {row}")

wb.close()

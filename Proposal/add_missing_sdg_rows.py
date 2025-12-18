"""
Add missing SDG rows to OQI Impact Workbook
Based on food production optimization context and relevant SDGs
"""

import openpyxl
from openpyxl.styles import Font, Alignment
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Load the workbook
wb_path = r"d:\Projects\OQI-UC002-DWave\Proposal\OQI_Impact_Workbook.xlsx"
wb = openpyxl.load_workbook(wb_path)
ws = wb['Impact Rows']

# Define SDG interlinkages for new rows
sdg_interlinkages = {
    'SDG 1': {
        'positive': 'SDG 2: Food security reduces poverty; SDG 8: Economic growth lifts people out of poverty',
        'negative': 'SDG 13: Climate policies may increase costs for poor households'
    },
    'SDG 5': {
        'positive': 'SDG 2: Women\'s empowerment improves food security; SDG 8: Gender equality boosts economic growth',
        'negative': 'SDG 9: Technology access gaps may worsen gender disparities'
    },
    'SDG 6': {
        'positive': 'SDG 3: Clean water improves health outcomes; SDG 15: Water management protects ecosystems',
        'negative': 'SDG 2: Agricultural water use may compete with drinking water needs'
    },
    'SDG 8': {
        'positive': 'SDG 1: Economic growth reduces poverty; SDG 9: Innovation drives job creation',
        'negative': 'SDG 13: Economic growth may increase emissions without green policies'
    },
    'SDG 10': {
        'positive': 'SDG 1: Reduced inequality lifts vulnerable populations; SDG 8: Fair wages support economic growth',
        'negative': 'SDG 9: Technology adoption may widen gaps without inclusive policies'
    },
    'SDG 12': {
        'positive': 'SDG 13: Responsible consumption reduces emissions; SDG 15: Reduced waste protects ecosystems',
        'negative': 'SDG 8: Consumption limits may slow short-term economic growth'
    },
    'SDG 15': {
        'positive': 'SDG 13: Ecosystem protection supports climate action; SDG 6: Forests regulate water cycles',
        'negative': 'SDG 2: Land conservation may limit agricultural expansion'
    },
    'SDG 17': {
        'positive': 'SDG 9: Partnerships enable technology transfer; SDG 2: Collaboration improves food systems',
        'negative': 'SDG 10: Unequal partnerships may reinforce power imbalances'
    }
}

# New rows to add (comprehensive coverage of relevant SDGs)
new_rows = [
    # SDG 1 - No Poverty
    {
        'ObjectiveLevel': 'SHORT_TERM_IMPACT',
        'ResultStatement': 'Increase incomes of small-scale farmers',
        'Indicator': '% increase in average income from optimized crop sales',
        'IndicatorDefinition': 'Change in farmer income from selling optimized high-value crops',
        'MeansOfMeasurement': 'Household income surveys comparing baseline to post-implementation',
        'Baseline': 'baseline surveys',
        'SDG': '1 No Poverty',
        'SDGTargets': '1.2 By 2030, reduce at least by half proportion of people living in poverty',
        'Positive': sdg_interlinkages['SDG 1']['positive'],
        'Negative': sdg_interlinkages['SDG 1']['negative']
    },
    {
        'ObjectiveLevel': 'LONG_TERM_IMPACT',
        'ResultStatement': 'Reduce poverty rates in farming communities',
        'Indicator': '% households above poverty line',
        'IndicatorDefinition': 'Share of farming households earning above national poverty threshold',
        'MeansOfMeasurement': 'Economic surveys at baseline and endline comparing income to poverty line',
        'Baseline': 'baseline surveys',
        'SDG': '1 No Poverty',
        'SDGTargets': '1.1 By 2030, eradicate extreme poverty for all people everywhere',
        'Positive': sdg_interlinkages['SDG 1']['positive'],
        'Negative': sdg_interlinkages['SDG 1']['negative']
    },
    
    # SDG 5 - Gender Equality
    {
        'ObjectiveLevel': 'SHORT_TERM_IMPACT',
        'ResultStatement': 'Increase women\'s participation in agricultural decision-making',
        'Indicator': '% women involved in crop planning decisions',
        'IndicatorDefinition': 'Share of households where women participate in selecting crops to plant',
        'MeansOfMeasurement': 'Gender-disaggregated surveys on farm decision-making processes',
        'Baseline': 'baseline surveys',
        'SDG': '5 Gender Equality',
        'SDGTargets': '5.a Reform to give women equal rights to economic resources & access to land',
        'Positive': sdg_interlinkages['SDG 5']['positive'],
        'Negative': sdg_interlinkages['SDG 5']['negative']
    },
    {
        'ObjectiveLevel': 'MID_TERM_IMPACT',
        'ResultStatement': 'Improve women\'s economic empowerment through agriculture',
        'Indicator': '% women with control over agricultural income',
        'IndicatorDefinition': 'Share of women who make decisions about spending farm income',
        'MeansOfMeasurement': 'Gender-disaggregated household surveys on income control',
        'Baseline': 'baseline surveys',
        'SDG': '5 Gender Equality',
        'SDGTargets': '5.5 Ensure full participation in leadership and decision-making',
        'Positive': sdg_interlinkages['SDG 5']['positive'],
        'Negative': sdg_interlinkages['SDG 5']['negative']
    },
    
    # SDG 6 - Clean Water
    {
        'ObjectiveLevel': 'OUTPUT',
        'ResultStatement': 'Optimize crop plans for water efficiency',
        'Indicator': 'Water use per kg of food produced in optimized plan',
        'IndicatorDefinition': 'Liters of water required per kilogram of nutritious food output',
        'MeansOfMeasurement': 'Water footprint analysis of optimized vs baseline crop selections',
        'Baseline': '',
        'SDG': '6 Clean Water and Sanitation',
        'SDGTargets': '6.4 By 2030, substantially increase water-use efficiency across all sectors',
        'Positive': sdg_interlinkages['SDG 6']['positive'],
        'Negative': sdg_interlinkages['SDG 6']['negative']
    },
    {
        'ObjectiveLevel': 'MID_TERM_IMPACT',
        'ResultStatement': 'Reduce agricultural water consumption',
        'Indicator': '% reduction in water use for food production',
        'IndicatorDefinition': 'Decrease in total water consumed for crop production per season',
        'MeansOfMeasurement': 'Water meter readings and irrigation monitoring systems',
        'Baseline': 'baseline surveys',
        'SDG': '6 Clean Water and Sanitation',
        'SDGTargets': '6.4 Substantially increase water-use efficiency',
        'Positive': sdg_interlinkages['SDG 6']['positive'],
        'Negative': sdg_interlinkages['SDG 6']['negative']
    },
    
    # SDG 8 - Decent Work and Economic Growth
    {
        'ObjectiveLevel': 'SHORT_TERM_IMPACT',
        'ResultStatement': 'Create employment in sustainable agriculture',
        'Indicator': 'Number of jobs created in optimized farming operations',
        'IndicatorDefinition': 'Additional labor positions generated by diversified crop production',
        'MeansOfMeasurement': 'Employment surveys before and after optimization implementation',
        'Baseline': 'baseline surveys',
        'SDG': '8 Decent Work and Economic Growth',
        'SDGTargets': '8.5 By 2030, achieve full employment and decent work with equal pay',
        'Positive': sdg_interlinkages['SDG 8']['positive'],
        'Negative': sdg_interlinkages['SDG 8']['negative']
    },
    {
        'ObjectiveLevel': 'LONG_TERM_IMPACT',
        'ResultStatement': 'Increase agricultural productivity and economic growth',
        'Indicator': 'GDP contribution from agricultural sector (%)',
        'IndicatorDefinition': 'Change in agricultural sector contribution to regional economy',
        'MeansOfMeasurement': 'Economic impact assessment using regional GDP data',
        'Baseline': 'baseline surveys',
        'SDG': '8 Decent Work and Economic Growth',
        'SDGTargets': '8.1 Sustain per capita economic growth',
        'Positive': sdg_interlinkages['SDG 8']['positive'],
        'Negative': sdg_interlinkages['SDG 8']['negative']
    },
    
    # SDG 10 - Reduced Inequalities
    {
        'ObjectiveLevel': 'SHORT_TERM_IMPACT',
        'ResultStatement': 'Ensure equitable access to optimization platform',
        'Indicator': '% of smallholder farmers with access to quantum planning tools',
        'IndicatorDefinition': 'Share of small-scale farmers able to use the optimization platform',
        'MeansOfMeasurement': 'Platform usage logs and farmer registration data',
        'Baseline': 'baseline surveys',
        'SDG': '10 Reduced Inequalities',
        'SDGTargets': '10.2 By 2030, empower and promote inclusion of all',
        'Positive': sdg_interlinkages['SDG 10']['positive'],
        'Negative': sdg_interlinkages['SDG 10']['negative']
    },
    {
        'ObjectiveLevel': 'MID_TERM_IMPACT',
        'ResultStatement': 'Reduce income inequality among farmers',
        'Indicator': 'Gini coefficient of farmer incomes',
        'IndicatorDefinition': 'Change in income distribution inequality measure',
        'MeansOfMeasurement': 'Income surveys calculating Gini coefficient at baseline and endline',
        'Baseline': 'baseline surveys',
        'SDG': '10 Reduced Inequalities',
        'SDGTargets': '10.1 By 2030, progressively achieve and sustain income growth of bottom 40%',
        'Positive': sdg_interlinkages['SDG 10']['positive'],
        'Negative': sdg_interlinkages['SDG 10']['negative']
    },
    
    # SDG 12 - Responsible Consumption and Production
    {
        'ObjectiveLevel': 'OUTPUT',
        'ResultStatement': 'Minimize food waste in production planning',
        'Indicator': '% reduction in food waste in optimized crop allocation',
        'IndicatorDefinition': 'Decrease in post-harvest losses through better planning',
        'MeansOfMeasurement': 'Food waste audits comparing optimized vs traditional planning',
        'Baseline': '',
        'SDG': '12 Responsible Consumption and Production',
        'SDGTargets': '12.3 By 2030, halve per capita global food waste',
        'Positive': sdg_interlinkages['SDG 12']['positive'],
        'Negative': sdg_interlinkages['SDG 12']['negative']
    },
    {
        'ObjectiveLevel': 'MID_TERM_IMPACT',
        'ResultStatement': 'Adopt sustainable agricultural practices',
        'Indicator': '% farms using sustainable production methods',
        'IndicatorDefinition': 'Share of farms implementing environmentally-sound practices',
        'MeansOfMeasurement': 'Farm practice surveys and certification records',
        'Baseline': 'baseline surveys',
        'SDG': '12 Responsible Consumption and Production',
        'SDGTargets': '12.2 By 2030, achieve sustainable management of natural resources',
        'Positive': sdg_interlinkages['SDG 12']['positive'],
        'Negative': sdg_interlinkages['SDG 12']['negative']
    },
    
    # SDG 15 - Life on Land
    {
        'ObjectiveLevel': 'MID_TERM_IMPACT',
        'ResultStatement': 'Protect biodiversity through sustainable land use',
        'Indicator': 'Biodiversity index on farmland',
        'IndicatorDefinition': 'Species richness and abundance on farms using optimized planning',
        'MeansOfMeasurement': 'Ecological surveys measuring species diversity',
        'Baseline': 'baseline surveys',
        'SDG': '15 Life on Land',
        'SDGTargets': '15.5 Take urgent action to reduce degradation and halt biodiversity loss',
        'Positive': sdg_interlinkages['SDG 15']['positive'],
        'Negative': sdg_interlinkages['SDG 15']['negative']
    },
    {
        'ObjectiveLevel': 'LONG_TERM_IMPACT',
        'ResultStatement': 'Prevent land degradation through optimized agriculture',
        'Indicator': 'Soil quality index',
        'IndicatorDefinition': 'Composite measure of soil health (organic matter, nutrients, pH)',
        'MeansOfMeasurement': 'Soil testing at baseline and regular intervals',
        'Baseline': 'baseline surveys',
        'SDG': '15 Life on Land',
        'SDGTargets': '15.3 By 2030, combat desertification and restore degraded land',
        'Positive': sdg_interlinkages['SDG 15']['positive'],
        'Negative': sdg_interlinkages['SDG 15']['negative']
    },
    
    # SDG 17 - Partnerships
    {
        'ObjectiveLevel': 'OUTPUT',
        'ResultStatement': 'Establish partnerships for technology deployment',
        'Indicator': 'Number of partnerships formed for quantum platform deployment',
        'IndicatorDefinition': 'Count of collaborations between research, government, NGOs, farmers',
        'MeansOfMeasurement': 'Partnership agreements and MOUs signed',
        'Baseline': '',
        'SDG': '17 Partnerships for the Goals',
        'SDGTargets': '17.16 Enhance global partnership for sustainable development',
        'Positive': sdg_interlinkages['SDG 17']['positive'],
        'Negative': sdg_interlinkages['SDG 17']['negative']
    },
    {
        'ObjectiveLevel': 'SHORT_TERM_IMPACT',
        'ResultStatement': 'Enable technology transfer and capacity building',
        'Indicator': 'Number of farmers/extension workers trained on platform',
        'IndicatorDefinition': 'Count of individuals receiving training on quantum optimization tools',
        'MeansOfMeasurement': 'Training attendance records and certification completion',
        'Baseline': 'baseline surveys',
        'SDG': '17 Partnerships for the Goals',
        'SDGTargets': '17.6 Enhance cooperation on science, technology and innovation',
        'Positive': sdg_interlinkages['SDG 17']['positive'],
        'Negative': sdg_interlinkages['SDG 17']['negative']
    },
]

# Get current row count
current_max_row = ws.max_row
print(f"Current row count: {current_max_row}")

# Add new rows
new_row_num = current_max_row + 1
rows_added = 0

for row_data in new_rows:
    ws.cell(row=new_row_num, column=1, value=row_data['ObjectiveLevel'])
    ws.cell(row=new_row_num, column=2, value=row_data['ResultStatement'])
    ws.cell(row=new_row_num, column=3, value=row_data['Indicator'])
    ws.cell(row=new_row_num, column=4, value=row_data['IndicatorDefinition'])
    ws.cell(row=new_row_num, column=5, value=row_data['MeansOfMeasurement'])
    ws.cell(row=new_row_num, column=6, value=row_data['Baseline'])
    ws.cell(row=new_row_num, column=7, value=row_data['SDG'])
    ws.cell(row=new_row_num, column=8, value=row_data['SDGTargets'])
    ws.cell(row=new_row_num, column=9, value=row_data['Positive'])
    ws.cell(row=new_row_num, column=10, value=row_data['Negative'])
    
    # Apply text wrapping
    for col in range(1, 11):
        ws.cell(row=new_row_num, column=col).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    print(f"Added row {new_row_num}: {row_data['SDG']} - {row_data['ResultStatement'][:60]}...")
    new_row_num += 1
    rows_added += 1

# Save the workbook
wb.save(wb_path)
print(f"\n✓ Added {rows_added} new rows to the workbook")
print(f"✓ Total rows now: {ws.max_row - 1}")

# Summary of SDG coverage
from collections import Counter
sdg_counts = Counter()
for row_idx in range(2, ws.max_row + 1):
    sdg_value = ws.cell(row=row_idx, column=7).value
    if sdg_value:
        sdg_num = sdg_value.split()[0]
        sdg_counts[sdg_num] += 1

print("\n=== SDG COVERAGE SUMMARY ===")
for sdg_num in sorted(sdg_counts.keys(), key=lambda x: int(x)):
    print(f"  SDG {sdg_num}: {sdg_counts[sdg_num]} rows")
